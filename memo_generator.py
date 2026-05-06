# ================================
# IMPORTS
# ================================
# os: clears the terminal screen on startup
# Path: cleaner file path handling than plain strings
# openpyxl: reads the allocations.xlsx file produced by allocator.py
# reportlab: generates the branded PDF memos
#   - SimpleDocTemplate: high-level layout engine (handles margins, pagination)
#   - Paragraph: a block of styled text
#   - Spacer: adds vertical whitespace between elements
#   - Table / TableStyle: draws formatted tables inside the PDF
#   - getSampleStyleSheet / ParagraphStyle: built-in and custom text styles
#   - colors / HRFlowable / letter: color palette, horizontal rules, page size
import os
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
)


# ================================
# PATH SETUP
# ================================
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "capital_calls" / "output"
MEMOS_DIR = OUTPUT_DIR / "memos"
ALLOCATIONS_FILE = OUTPUT_DIR / "allocations.xlsx"


# ================================
# FIRM CONSTANTS
# ================================
# These appear on every memo — update here if branding changes
FIRM_NAME = "HighTraverse Capital Operations"
FIRM_EMAIL = "info@hightraverse.co"
FIRM_WEBSITE = "hightraverse.co"


# ================================
# STYLES
# ================================
def build_styles():
    """
    Returns a dict of named ParagraphStyles used throughout the memo.
    We define our own on top of the default stylesheet so fonts and
    sizes are consistent across every generated PDF.
    """
    base = getSampleStyleSheet()

    styles = {
        # Firm name at the top of the letterhead
        "firm_name": ParagraphStyle(
            "firm_name",
            parent=base["Normal"],
            fontSize=16,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1a2e4a"),  # dark navy
            spaceAfter=2,
        ),
        # Firm contact line under the firm name
        "firm_contact": ParagraphStyle(
            "firm_contact",
            parent=base["Normal"],
            fontSize=9,
            fontName="Helvetica",
            textColor=colors.HexColor("#555555"),
            spaceAfter=0,
        ),
        # "CAPITAL CALL NOTICE" label
        "doc_type": ParagraphStyle(
            "doc_type",
            parent=base["Normal"],
            fontSize=10,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1a2e4a"),
            spaceBefore=18,
            spaceAfter=4,
        ),
        # Section labels ("TO:", "RE:", etc.)
        "label": ParagraphStyle(
            "label",
            parent=base["Normal"],
            fontSize=10,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#333333"),
            spaceAfter=2,
        ),
        # Normal body text
        "body": ParagraphStyle(
            "body",
            parent=base["Normal"],
            fontSize=10,
            fontName="Helvetica",
            textColor=colors.HexColor("#333333"),
            leading=15,       # line height
            spaceAfter=8,
        ),
        # The allocation amount shown in large type
        "amount_label": ParagraphStyle(
            "amount_label",
            parent=base["Normal"],
            fontSize=10,
            fontName="Helvetica",
            textColor=colors.HexColor("#555555"),
            spaceAfter=2,
        ),
        "amount_value": ParagraphStyle(
            "amount_value",
            parent=base["Normal"],
            fontSize=22,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1a2e4a"),
            spaceAfter=6,
        ),
        # Footer text
        "footer": ParagraphStyle(
            "footer",
            parent=base["Normal"],
            fontSize=8,
            fontName="Helvetica",
            textColor=colors.HexColor("#888888"),
        ),
    }
    return styles


# ================================
# READ ALLOCATIONS FILE
# ================================
def read_allocations(file_path: Path) -> dict:
    """
    Reads allocations.xlsx produced by allocator.py.

    The file has a 5-row header block followed by column headers on row 6
    and LP data from row 7 onward. We extract both the deal metadata
    (investment name, vehicle, due date, total call) and the per-LP rows.

    Returns a dict with keys:
        investment_name, vehicle_name, due_date, total_call, investors
    where investors is a list of (investor_name, ownership_pct, allocation) tuples.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"allocations.xlsx not found at: {file_path}\n"
            "Run allocator.py first to generate the allocations file."
        )

    wb = load_workbook(file_path)
    ws = wb["Allocations"]

    # Rows 1-5 are the header block written by format_excel() in allocator.py
    # Each row is (label, value, None, None)
    investment_name = ws["B2"].value or "Unknown Investment"
    vehicle_name    = ws["B3"].value or "Unknown Vehicle"
    due_date        = ws["B4"].value or "Not Found"
    total_call      = ws["B5"].value or 0.0

    # Row 6 = column headers; data starts at row 7
    investors = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        # Skip blank rows and the TOTAL summary row at the bottom
        if not row[1] or str(row[1]).strip().upper() == "TOTAL":
            continue
        investor_name  = str(row[1]).strip()
        ownership_pct  = float(row[2]) if row[2] is not None else 0.0
        allocation     = float(row[3]) if row[3] is not None else 0.0
        investors.append((investor_name, ownership_pct, allocation))

    print(f"  Investment  : {investment_name}")
    print(f"  Vehicle     : {vehicle_name}")
    print(f"  Due date    : {due_date}")
    print(f"  Total call  : ${total_call:,.2f}")
    print(f"  Investors   : {len(investors)}")

    return {
        "investment_name": investment_name,
        "vehicle_name":    vehicle_name,
        "due_date":        due_date,
        "total_call":      total_call,
        "investors":       investors,
    }


# ================================
# GENERATE ONE MEMO
# ================================
def generate_memo(investor_name: str, ownership_pct: float, allocation: float,
                  deal: dict, styles: dict, output_path: Path):
    """
    Generates a single branded capital call memo PDF for one LP.

    Layout:
      - Letterhead (firm name + contact)
      - Horizontal rule
      - Document type label
      - Memo header table (TO / FROM / DATE / RE)
      - Body paragraph explaining the capital call
      - Allocation summary box (their specific amount)
      - Details table (investment, vehicle, due date, their %)
      - Wire instruction placeholder
      - Footer

    Parameters
    ----------
    investor_name  : LP name as it appears in vehicle_investors table
    ownership_pct  : decimal ownership (e.g. 0.125 = 12.5%)
    allocation     : dollar amount owed by this LP
    deal           : dict returned by read_allocations()
    styles         : dict returned by build_styles()
    output_path    : where to write the PDF file
    """
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        leftMargin=0.9 * inch,
        rightMargin=0.9 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.8 * inch,
    )

    story = []  # story = ordered list of flowable elements that make up the page

    # ------------------------------------------------------------------
    # LETTERHEAD
    # ------------------------------------------------------------------
    story.append(Paragraph(FIRM_NAME, styles["firm_name"]))
    story.append(Paragraph(
        f"{FIRM_EMAIL}  ·  {FIRM_WEBSITE}",
        styles["firm_contact"],
    ))

    # Navy horizontal rule under the letterhead
    story.append(Spacer(1, 6))
    story.append(HRFlowable(
        width="100%",
        thickness=2,
        color=colors.HexColor("#1a2e4a"),
        spaceAfter=0,
    ))

    # ------------------------------------------------------------------
    # DOCUMENT TYPE LABEL
    # ------------------------------------------------------------------
    story.append(Paragraph("CAPITAL CALL NOTICE", styles["doc_type"]))

    # ------------------------------------------------------------------
    # MEMO HEADER TABLE (TO / FROM / DATE / RE)
    # ------------------------------------------------------------------
    # We use a two-column table so labels and values stay aligned cleanly
    from datetime import date
    today = date.today().strftime("%B %d, %Y")

    header_data = [
        [Paragraph("<b>TO:</b>",   styles["body"]), Paragraph(investor_name,                  styles["body"])],
        [Paragraph("<b>FROM:</b>", styles["body"]), Paragraph(FIRM_NAME,                      styles["body"])],
        [Paragraph("<b>DATE:</b>", styles["body"]), Paragraph(today,                           styles["body"])],
        [Paragraph("<b>RE:</b>",   styles["body"]), Paragraph(
            f"Capital Call — {deal['investment_name']} / {deal['vehicle_name']}",
            styles["body"],
        )],
    ]

    header_table = Table(header_data, colWidths=[0.8 * inch, 5.8 * inch])
    header_table.setStyle(TableStyle([
        ("VALIGN",    (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(header_table)

    # Light rule separating header from body
    story.append(HRFlowable(
        width="100%",
        thickness=0.5,
        color=colors.HexColor("#cccccc"),
        spaceBefore=4,
        spaceAfter=14,
    ))

    # ------------------------------------------------------------------
    # BODY PARAGRAPH
    # ------------------------------------------------------------------
    body_text = (
        f"Please be advised that {deal['investment_name']} (the \"Fund\") has issued a capital "
        f"call notice to {deal['vehicle_name']} (the \"Vehicle\"). Pursuant to your limited "
        f"partnership interest in the Vehicle, your pro-rata capital contribution is set forth below."
        f"<br/><br/>"
        f"Payment is due no later than <b>{deal['due_date']}</b>. Wire instructions are provided "
        f"at the bottom of this notice. Please ensure funds are received by the due date to avoid "
        f"any default provisions under the limited partnership agreement."
    )
    story.append(Paragraph(body_text, styles["body"]))

    # ------------------------------------------------------------------
    # ALLOCATION AMOUNT BOX
    # ------------------------------------------------------------------
    # A shaded box that highlights the LP's specific dollar amount
    story.append(Spacer(1, 6))
    story.append(Paragraph("YOUR CAPITAL CALL AMOUNT", styles["amount_label"]))
    story.append(Paragraph(f"${allocation:,.2f}", styles["amount_value"]))

    story.append(HRFlowable(
        width="100%",
        thickness=0.5,
        color=colors.HexColor("#cccccc"),
        spaceBefore=2,
        spaceAfter=14,
    ))

    # ------------------------------------------------------------------
    # DETAILS TABLE
    # ------------------------------------------------------------------
    details_data = [
        ["Fund / Investment",   deal["investment_name"]],
        ["Vehicle",             deal["vehicle_name"]],
        ["Total Capital Call",  f"${deal['total_call']:,.2f}"],
        ["Your Ownership %",    f"{ownership_pct:.4%}"],
        ["Your Allocation",     f"${allocation:,.2f}"],
        ["Payment Due Date",    str(deal["due_date"])],
    ]

    details_table = Table(details_data, colWidths=[2.2 * inch, 4.4 * inch])
    details_table.setStyle(TableStyle([
        # Header column styling
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("TEXTCOLOR",   (0, 0), (0, -1), colors.HexColor("#333333")),
        ("TEXTCOLOR",   (1, 0), (1, -1), colors.HexColor("#555555")),
        # Alternating row shading for readability
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#f0f4f8")),
        ("BACKGROUND",  (0, 2), (-1, 2), colors.HexColor("#f0f4f8")),
        ("BACKGROUND",  (0, 4), (-1, 4), colors.HexColor("#f0f4f8")),
        # Highlight the allocation row
        ("BACKGROUND",  (0, 4), (-1, 4), colors.HexColor("#d5e8f0")),
        ("FONTNAME",    (0, 4), (-1, 4), "Helvetica-Bold"),
        # Grid and padding
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(details_table)

    # ------------------------------------------------------------------
    # WIRE INSTRUCTIONS PLACEHOLDER
    # ------------------------------------------------------------------
    story.append(Spacer(1, 20))
    story.append(Paragraph("WIRE INSTRUCTIONS", styles["label"]))
    story.append(Paragraph(
        "Wire instructions will be provided by the fund administrator. "
        "Please contact info@hightraverse.co if you have not received wire details.",
        styles["body"],
    ))

    # ------------------------------------------------------------------
    # FOOTER
    # ------------------------------------------------------------------
    story.append(Spacer(1, 30))
    story.append(HRFlowable(
        width="100%",
        thickness=0.5,
        color=colors.HexColor("#cccccc"),
        spaceAfter=6,
    ))
    story.append(Paragraph(
        f"This notice has been prepared by {FIRM_NAME} on behalf of {deal['vehicle_name']}. "
        f"This document is confidential and intended solely for the named recipient. "
        f"Questions: {FIRM_EMAIL}",
        styles["footer"],
    ))

    # Build (render) the PDF
    doc.build(story)


# ================================
# MAIN EXECUTION
# ================================
def main():
    os.system("cls" if os.name == "nt" else "clear")

    print("=" * 50)
    print("  Capital Call Memo Generator")
    print("=" * 50)
    print()

    # Step 1: Read allocations file
    print("Step 1: Reading allocations file...")
    deal = read_allocations(ALLOCATIONS_FILE)
    print()

    # Step 2: Set up output folder
    print("Step 2: Setting up output folder...")
    MEMOS_DIR.mkdir(parents=True, exist_ok=True)
    print(f"  Memos folder: {MEMOS_DIR}")
    print()

    # Step 3: Build styles once, reuse for every memo
    styles = build_styles()

    # Step 4: Generate one PDF per LP
    print("Step 3: Generating memos...")
    generated = []
    errors = []

    for investor_name, ownership_pct, allocation in deal["investors"]:
        # Build a safe filename: replace any characters that aren't
        # letters, numbers, spaces, or hyphens with underscores
        safe_name = "".join(
            c if c.isalnum() or c in (" ", "-") else "_"
            for c in investor_name
        ).strip()
        output_path = MEMOS_DIR / f"{safe_name} - Capital Call Memo.pdf"

        try:
            generate_memo(investor_name, ownership_pct, allocation, deal, styles, output_path)
            print(f"  [OK] {output_path.name}")
            generated.append(output_path)
        except Exception as e:
            print(f"  [ERROR] {investor_name}: {e}")
            errors.append((investor_name, str(e)))

    print()

    # Step 5: Summary
    print("=" * 50)
    print("  COMPLETE")
    print("=" * 50)
    print(f"  Memos generated : {len(generated)}")
    if errors:
        print(f"  Errors          : {len(errors)}")
        for name, err in errors:
            print(f"    - {name}: {err}")
    print(f"  Output folder   : {MEMOS_DIR}")
    print()


# ================================
# RUN SCRIPT
# ================================
if __name__ == "__main__":
    main()
