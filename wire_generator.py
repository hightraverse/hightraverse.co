# ================================
# IMPORTS
# ================================
# os: clears the terminal on startup
# re: regex for extracting wire fields from PDF text
# json: serializes data for Supabase POST requests
# datetime: timestamps new wire instruction records
# Path: cleaner file path handling
# pdfplumber: extracts text from the capital call PDF
# requests: hits the Supabase REST API (read and write)
# load_dotenv: loads SUPABASE_URL and SUPABASE_KEY from .env
# openpyxl: reads allocations.xlsx for deal metadata
# reportlab: generates the wire template PDF
import os
import re
import json
from datetime import date
from pathlib import Path

import pdfplumber
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
)


# ================================
# PATH SETUP
# ================================
BASE_DIR    = Path(__file__).resolve().parent
DATA_DIR    = BASE_DIR / "capital_calls" / "data"
OUTPUT_DIR  = BASE_DIR / "capital_calls" / "output"
PDF_PATH    = DATA_DIR / "sample.pdf"
ALLOC_FILE  = OUTPUT_DIR / "allocations.xlsx"


# ================================
# FIRM CONSTANTS
# ================================
FIRM_NAME    = "HighTraverse Capital Operations"
FIRM_EMAIL   = "info@hightraverse.co"


# ================================
# SUPABASE CONNECTION
# ================================
def get_supabase_client():
    """
    Loads credentials from .env and returns a (base_url, headers) tuple.
    The headers work for both GET and POST requests to Supabase REST API.
    """
    load_dotenv(BASE_DIR / ".env")
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        raise ValueError("SUPABASE_URL or SUPABASE_KEY not found in .env")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    return url.rstrip("/"), headers


def fetch_table(client, table_name: str, filters: dict = None) -> list:
    """
    GET all rows from a Supabase table, with optional column=value filters.
    filters example: {"InvestmentName": "ABC Fund, LP"}
    """
    base_url, headers = client
    params = {"select": "*"}
    if filters:
        # Supabase REST filter syntax: column=eq.value
        for col, val in filters.items():
            params[col] = f"eq.{val}"
    response = requests.get(
        f"{base_url}/rest/v1/{table_name}",
        headers=headers,
        params=params,
        timeout=15,
    )
    response.raise_for_status()
    return response.json()


def insert_row(client, table_name: str, data: dict) -> dict:
    """
    POST a new row into a Supabase table.
    Returns the inserted row (Supabase echoes it back with its new id).
    """
    base_url, headers = client
    # Prefer: return=representation tells Supabase to return the inserted row
    post_headers = {**headers, "Prefer": "return=representation"}
    response = requests.post(
        f"{base_url}/rest/v1/{table_name}",
        headers=post_headers,
        data=json.dumps(data),
        timeout=15,
    )
    response.raise_for_status()
    result = response.json()
    # Supabase returns a list with the inserted row
    return result[0] if result else {}


# ================================
# READ ALLOCATIONS FILE
# ================================
def read_allocations(file_path: Path) -> dict:
    """
    Reads the header block from allocations.xlsx (rows 1-5) to get
    investment name, vehicle name, due date, and total call amount.
    We don't need the LP rows here — just the deal-level metadata.
    """
    if not file_path.exists():
        raise FileNotFoundError(
            f"allocations.xlsx not found at: {file_path}\n"
            "Run allocator.py first."
        )
    wb = load_workbook(file_path)
    ws = wb["Allocations"]
    return {
        "investment_name": ws["B2"].value or "Unknown",
        "vehicle_name":    ws["B3"].value or "Unknown",
        "due_date":        str(ws["B4"].value or "Not Found"),
        "total_call":      float(ws["B5"].value or 0.0),
    }


# ================================
# EXTRACT WIRE INSTRUCTIONS FROM PDF
# ================================
def extract_wire_instructions(pdf_path: Path) -> dict:
    """
    Reads the capital call PDF and uses regex to pull out wire instruction fields.

    We look for labelled lines in the format "Field Name: value".
    Each regex is anchored to its label so we don't accidentally pick up
    unrelated numbers (e.g. confusing a date with an ABA routing number).

    Returns a dict with keys matching the received_wire_instructions table columns.
    Any field not found is returned as an empty string so callers can check cleanly.
    """
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

    def extract(pattern):
        """Helper: run a regex and return the first captured group, or ''."""
        match = re.search(pattern, text, re.IGNORECASE)
        return match.group(1).strip() if match else ""

    wire = {
        "BankName":        extract(r"Bank Name:\s*(.+)"),
        "ABARouting":      extract(r"ABA Routing(?:\s+Number)?:\s*([\d\-]+)"),
        "AccountNumber":   extract(r"Account Number:\s*([\w\-]+)"),
        "BeneficiaryName": extract(r"Account Name:\s*(.+)"),
        "SwiftCode":       extract(r"Swift(?:\s+Code)?:\s*([\w]+)"),
        "BankAddress":     extract(r"Bank Address:\s*(.+)"),
        "Reference":       extract(r"Reference:\s*(.+)"),
    }

    return wire


# ================================
# WIRE INSTRUCTION COMPARISON
# ================================
# Status codes returned by compare_wire_instructions()
STATUS_NEW     = "NEW"       # first time we've seen this fund's wire details
STATUS_MATCH   = "MATCH"     # identical to previous call — safe to proceed
STATUS_CHANGED = "CHANGED"   # critical fields differ — verbal confirmation required

# These are the fields that must match exactly. A change in any of these
# is a red flag for potential wire fraud and requires verbal confirmation.
CRITICAL_FIELDS = ["ABARouting", "AccountNumber", "BeneficiaryName"]


def compare_wire_instructions(new_wire: dict, history: list) -> tuple:
    """
    Compares the freshly extracted wire instructions against every previous
    record for this fund stored in received_wire_instructions.

    We compare only the CRITICAL_FIELDS. Non-critical differences (e.g. a
    reference line that changed) are noted but don't require verbal confirmation.

    Returns:
        (status, changed_fields)
        status        : STATUS_NEW, STATUS_MATCH, or STATUS_CHANGED
        changed_fields: list of field names that differ (empty if MATCH or NEW)
    """
    if not history:
        return STATUS_NEW, []

    # Use the most recent record for comparison (highest id = most recent insert)
    latest = max(history, key=lambda r: r.get("id", 0))

    changed = []
    for field in CRITICAL_FIELDS:
        prev_val = str(latest.get(field, "")).strip()
        new_val  = str(new_wire.get(field, "")).strip()
        if prev_val and new_val and prev_val != new_val:
            changed.append(field)

    if changed:
        return STATUS_CHANGED, changed
    return STATUS_MATCH, []


# ================================
# REPORTLAB STYLES
# ================================
def build_styles():
    """Returns named ParagraphStyles used in the wire template PDF."""
    base = getSampleStyleSheet()
    navy = colors.HexColor("#1a2e4a")
    grey = colors.HexColor("#555555")

    return {
        "firm_name": ParagraphStyle("firm_name", parent=base["Normal"],
            fontSize=16, fontName="Helvetica-Bold", textColor=navy, spaceAfter=2),
        "firm_contact": ParagraphStyle("firm_contact", parent=base["Normal"],
            fontSize=9, fontName="Helvetica", textColor=grey, spaceAfter=0),
        "doc_type": ParagraphStyle("doc_type", parent=base["Normal"],
            fontSize=10, fontName="Helvetica-Bold", textColor=navy,
            spaceBefore=16, spaceAfter=4),
        "section_header": ParagraphStyle("section_header", parent=base["Normal"],
            fontSize=9, fontName="Helvetica-Bold", textColor=navy,
            spaceBefore=12, spaceAfter=4),
        "body": ParagraphStyle("body", parent=base["Normal"],
            fontSize=10, fontName="Helvetica", textColor=colors.HexColor("#333333"),
            leading=15, spaceAfter=6),
        "alert_text": ParagraphStyle("alert_text", parent=base["Normal"],
            fontSize=10, fontName="Helvetica-Bold",
            textColor=colors.HexColor("#cc0000"), spaceAfter=4),
        "footer": ParagraphStyle("footer", parent=base["Normal"],
            fontSize=8, fontName="Helvetica", textColor=grey),
    }


# ================================
# GENERATE WIRE TEMPLATE PDF
# ================================
def generate_wire_pdf(deal: dict, new_wire: dict, standing: dict,
                      status: str, changed_fields: list,
                      output_path: Path, styles: dict):
    """
    Generates the wire instruction template PDF.

    Sections:
      1. Letterhead
      2. Status alert box (NEW / MATCH / VERBAL CONFIRMATION REQUIRED)
      3. Wire details — Originator (our vehicle) and Beneficiary (the fund)
      4. Payment details (amount, reference, due date)
      5. Approval sign-off lines
      6. Footer
    """
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        leftMargin=0.9 * inch, rightMargin=0.9 * inch,
        topMargin=0.8 * inch,  bottomMargin=0.8 * inch,
    )
    story = []

    # ------------------------------------------------------------------
    # LETTERHEAD
    # ------------------------------------------------------------------
    story.append(Paragraph(FIRM_NAME, styles["firm_name"]))
    story.append(Paragraph(f"{FIRM_EMAIL}  ·  hightraverse.co", styles["firm_contact"]))
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=2,
        color=colors.HexColor("#1a2e4a"), spaceAfter=0))

    story.append(Paragraph("OUTBOUND WIRE INSTRUCTION TEMPLATE", styles["doc_type"]))
    story.append(Paragraph(
        f"Prepared: {date.today().strftime('%B %d, %Y')}  ·  "
        f"Due Date: {deal['due_date']}",
        styles["body"],
    ))

    # ------------------------------------------------------------------
    # STATUS ALERT BOX
    # ------------------------------------------------------------------
    # The status box is the most important element — it tells the ops
    # team whether they need to pick up the phone before sending this wire.
    if status == STATUS_CHANGED:
        alert_color = colors.HexColor("#fff0f0")   # light red
        border_color = colors.HexColor("#cc0000")
        alert_lines = [
            [Paragraph("⚠  VERBAL CONFIRMATION REQUIRED", styles["alert_text"])],
            [Paragraph(
                f"The following wire fields have changed since the last capital call from "
                f"{deal['investment_name']}: <b>{', '.join(changed_fields)}</b>.<br/>"
                "Do NOT send this wire until you have verbally confirmed the new banking "
                "details directly with the fund's investor relations team.",
                styles["body"],
            )],
        ]
    elif status == STATUS_NEW:
        alert_color = colors.HexColor("#fffbe6")   # light amber
        border_color = colors.HexColor("#e6a817")
        alert_lines = [
            [Paragraph("ℹ  NEW COUNTERPARTY — FIRST CAPITAL CALL", styles["alert_text"])],
            [Paragraph(
                f"No prior wire instructions on file for {deal['investment_name']}. "
                "These instructions have been stored for future comparison. "
                "Consider verbally confirming banking details before sending.",
                styles["body"],
            )],
        ]
    else:  # MATCH
        alert_color = colors.HexColor("#f0fff4")   # light green
        border_color = colors.HexColor("#2e7d32")
        alert_lines = [
            [Paragraph("✓  WIRE INSTRUCTIONS VERIFIED — MATCH ON FILE", styles["alert_text"])],
            [Paragraph(
                f"Wire instructions match the previously verified record for "
                f"{deal['investment_name']}. No verbal confirmation required.",
                styles["body"],
            )],
        ]

    alert_table = Table(alert_lines, colWidths=[6.4 * inch])
    alert_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), alert_color),
        ("LINEABOVE",     (0, 0), (-1, 0),  1.5, border_color),
        ("LINEBELOW",     (0, -1), (-1, -1), 1.5, border_color),
        ("LINEBEFORE",    (0, 0), (0, -1),  1.5, border_color),
        ("LINEAFTER",     (-1, 0), (-1, -1), 1.5, border_color),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
    ]))
    story.append(Spacer(1, 10))
    story.append(alert_table)

    # ------------------------------------------------------------------
    # WIRE DETAILS — two-column table per section
    # ------------------------------------------------------------------
    def wire_section(title: str, rows: list):
        """
        Helper that renders a labelled section table.
        rows = list of (label, value) tuples.
        """
        story.append(Paragraph(title, styles["section_header"]))
        data = [[Paragraph(f"<b>{lbl}</b>", styles["body"]),
                 Paragraph(str(val) if val else "—", styles["body"])]
                for lbl, val in rows]
        tbl = Table(data, colWidths=[2.0 * inch, 4.4 * inch])
        tbl.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("LINEBELOW",     (0, 0), (-1, -2), 0.5, colors.HexColor("#eeeeee")),
            ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ]))
        story.append(tbl)

    # Originator = our fund-of-funds vehicle sending the wire
    wire_section("ORIGINATOR (SENDING ACCOUNT)", [
        ("Account Name",    standing.get("AccountName")),
        ("Bank Name",       standing.get("BankName")),
        ("ABA Routing",     standing.get("ABARouting")),
        ("Account Number",  standing.get("AccountNumber")),
        ("Bank Address",    standing.get("BankAddress")),
    ])

    # Beneficiary = the underlying fund receiving the wire
    wire_section("BENEFICIARY (RECEIVING ACCOUNT)", [
        ("Account Name",    new_wire.get("BeneficiaryName")),
        ("Bank Name",       new_wire.get("BankName")),
        ("ABA Routing",     new_wire.get("ABARouting")),
        ("Account Number",  new_wire.get("AccountNumber")),
        ("Swift Code",      new_wire.get("SwiftCode") or "N/A"),
        ("Bank Address",    new_wire.get("BankAddress") or "N/A"),
    ])

    # Payment details
    wire_section("PAYMENT DETAILS", [
        ("Amount",          f"${deal['total_call']:,.2f}"),
        ("Value Date",      deal["due_date"]),
        ("Reference / Memo", new_wire.get("Reference")),
        ("Investment",      deal["investment_name"]),
        ("Vehicle",         deal["vehicle_name"]),
    ])

    # ------------------------------------------------------------------
    # APPROVAL SIGN-OFF
    # ------------------------------------------------------------------
    story.append(Spacer(1, 20))
    story.append(Paragraph("AUTHORIZATION", styles["section_header"]))

    signoff_data = [
        ["Prepared by:",        "_" * 35,  "Date:", "_" * 20],
        ["Reviewed by:",        "_" * 35,  "Date:", "_" * 20],
        ["Authorized by:",      "_" * 35,  "Date:", "_" * 20],
    ]
    if status == STATUS_CHANGED:
        signoff_data.append(
            ["Verbal Confirm (name):", "_" * 35, "Call time:", "_" * 20]
        )

    signoff_table = Table(signoff_data, colWidths=[1.5*inch, 2.8*inch, 0.7*inch, 1.7*inch])
    signoff_table.setStyle(TableStyle([
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TEXTCOLOR",     (0, 0), (-1, -1), colors.HexColor("#333333")),
    ]))
    story.append(signoff_table)

    # ------------------------------------------------------------------
    # FOOTER
    # ------------------------------------------------------------------
    story.append(Spacer(1, 24))
    story.append(HRFlowable(width="100%", thickness=0.5,
        color=colors.HexColor("#cccccc"), spaceAfter=6))
    story.append(Paragraph(
        f"Generated by {FIRM_NAME}. Confidential — for internal use only. "
        f"Do not share outside the organization. Questions: {FIRM_EMAIL}",
        styles["footer"],
    ))

    doc.build(story)


# ================================
# MAIN EXECUTION
# ================================
def main():
    os.system("cls" if os.name == "nt" else "clear")

    print("=" * 50)
    print("  Wire Template Generator")
    print("=" * 50)
    print()

    # Step 1: Connect to Supabase
    print("Step 1: Connecting to Supabase...")
    client = get_supabase_client()
    print("  Connected.")
    print()

    # Step 2: Read deal metadata from allocations.xlsx
    print("Step 2: Reading allocations file...")
    deal = read_allocations(ALLOC_FILE)
    print(f"  Investment : {deal['investment_name']}")
    print(f"  Vehicle    : {deal['vehicle_name']}")
    print(f"  Total call : ${deal['total_call']:,.2f}")
    print(f"  Due date   : {deal['due_date']}")
    print()

    # Step 3: Extract wire instructions from the capital call PDF
    print("Step 3: Extracting wire instructions from PDF...")
    new_wire = extract_wire_instructions(PDF_PATH)
    if not new_wire["ABARouting"] and not new_wire["AccountNumber"]:
        print("  [WARNING] No wire instructions found in the PDF.")
        print("  Check that the PDF contains lines like:")
        print("    ABA Routing Number: 026015079")
        print("    Account Number: 112233445")
        return
    print(f"  Bank         : {new_wire['BankName']}")
    print(f"  ABA Routing  : {new_wire['ABARouting']}")
    print(f"  Account      : {new_wire['AccountNumber']}")
    print(f"  Beneficiary  : {new_wire['BeneficiaryName']}")
    print(f"  Reference    : {new_wire['Reference']}")
    print()

    # Step 4: Look up wire instruction history for this fund
    print("Step 4: Checking wire instruction history...")
    history = fetch_table(
        client,
        "received_wire_instructions",
        filters={"InvestmentName": deal["investment_name"]},
    )
    print(f"  Previous records found: {len(history)}")

    status, changed_fields = compare_wire_instructions(new_wire, history)

    if status == STATUS_NEW:
        print("  Status: NEW — first capital call from this fund")
    elif status == STATUS_MATCH:
        print("  Status: MATCH — wire instructions unchanged")
    else:
        print(f"  Status: CHANGED — fields differ: {', '.join(changed_fields)}")
        print("  *** VERBAL CONFIRMATION REQUIRED ***")
    print()

    # Step 5: Store the extracted wire instructions in Supabase
    print("Step 5: Storing wire instructions in Supabase...")
    record = {
        "InvestmentName":    deal["investment_name"],
        "BankName":          new_wire["BankName"],
        "ABARouting":        new_wire["ABARouting"],
        "AccountNumber":     new_wire["AccountNumber"],
        "BeneficiaryName":   new_wire["BeneficiaryName"],
        "SwiftCode":         new_wire["SwiftCode"],
        "BankAddress":       new_wire["BankAddress"],
        "Reference":         new_wire["Reference"],
        "ExtractedDate":     date.today().isoformat(),
        "SourceFile":        PDF_PATH.name,
        "VerballyConfirmed": False,
    }
    inserted = insert_row(client, "received_wire_instructions", record)
    print(f"  Stored with id: {inserted.get('id')}")
    print()

    # Step 6: Load our vehicle's standing (outbound) wire instructions
    print("Step 6: Loading vehicle standing instructions...")
    standing_rows = fetch_table(
        client,
        "vehicle_standing_instructions",
        filters={"VehicleName": deal["vehicle_name"]},
    )
    if not standing_rows:
        print(f"  [ERROR] No standing instructions found for vehicle: {deal['vehicle_name']}")
        print("  Add a row to vehicle_standing_instructions in Supabase and re-run.")
        return
    standing = standing_rows[0]
    print(f"  Found: {standing['AccountName']} at {standing['BankName']}")
    print()

    # Step 7: Generate the wire template PDF
    print("Step 7: Generating wire template PDF...")
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / f"{deal['investment_name']} - Wire Template.pdf"
    styles = build_styles()
    generate_wire_pdf(deal, new_wire, standing, status, changed_fields, output_path, styles)
    print(f"  Saved to: {output_path}")
    print()

    # Summary
    print("=" * 50)
    print("  COMPLETE")
    print("=" * 50)
    print(f"  Status   : {status}")
    if changed_fields:
        print(f"  Changed  : {', '.join(changed_fields)}")
    print(f"  Output   : {output_path}")
    print()


# ================================
# RUN SCRIPT
# ================================
if __name__ == "__main__":
    main()
