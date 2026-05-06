# ================================
# IMPORTS
# ================================
# os: used to clear the terminal screen
# re: used for regex (pattern matching inside PDF text)
# Optional: lets us say "this might return a float OR None"
# Path: a cleaner way to work with file paths than plain strings
# pd: pandas - reads/writes Excel and CSV files, handles tables of data
# pdfplumber: opens PDF files and extracts the raw text from each page
# requests: makes HTTP calls to the Supabase REST API
# load_workbook / Font: opens the Excel file we created so we can apply formatting
# get_column_letter: converts a column number (4) to a letter ("D") dynamically
# load_dotenv: reads your .env file so Python can access SUPABASE_URL and SUPABASE_KEY
import os
import re
from typing import Optional
from pathlib import Path

import pandas as pd
import pdfplumber
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Why we use `requests` instead of the official `supabase` Python SDK:
# The supabase package pulls in `pyiceberg`, which requires Microsoft Visual
# C++ Build Tools to compile on Windows. Rather than install several GB of
# build tools, we hit Supabase's REST API directly. Same result, no
# compilation, no dependency hell.


# ================================
# PATH SETUP
# ================================
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "capital_calls" / "data"
OUTPUT_DIR = BASE_DIR / "capital_calls" / "output"
LOG_DIR = BASE_DIR / "capital_calls" / "logs"


# ================================
# SUPABASE CONNECTION
# ================================
def get_supabase_client():
    """
    Loads SUPABASE_URL and SUPABASE_KEY from .env, then returns a
    (base_url, headers) tuple used everywhere we hit Supabase.

    The headers contain:
      - apikey: the Supabase JWT (service_role)
      - Authorization: Bearer <same key> (REST API requires both)

    The .rstrip('/') just trims any accidental trailing slash on the URL
    so we never end up with double slashes when we build endpoints.
    """
    load_dotenv(BASE_DIR / ".env")

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")

    if not url or not key:
        raise ValueError("SUPABASE_URL or SUPABASE_KEY not found.")

    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
    }
    return url.rstrip("/"), headers


def fetch_table(client, table_name: str) -> list:
    """
    Fetches every row from a Supabase table via the REST API.

    The REST endpoint Supabase exposes for any table is:
        {url}/rest/v1/{table_name}?select=*

    We pass the auth headers, ask for all columns, and let `requests`
    raise a clear error if the call fails (raise_for_status). The
    .json() call returns a list of dicts — one dict per row — which
    is the same shape the supabase SDK used to return in response.data,
    so the rest of the script doesn't change.
    """
    base_url, headers = client
    endpoint = f"{base_url}/rest/v1/{table_name}"
    response = requests.get(
        endpoint,
        headers=headers,
        params={"select": "*"},
        timeout=15,
    )
    response.raise_for_status()
    return response.json()

# ================================
# LOGGING
# ================================
def log_run(investment: str, vehicle: str, total_call: float, investor_count: int, output_file: Path):
    """
    Appends one row to logs/run_log.csv each time the tool runs successfully.
    """
    LOG_DIR.mkdir(exist_ok=True)
    log_file = LOG_DIR / "run_log.csv"

    new_row = pd.DataFrame([{
        "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "InvestmentName": investment,
        "VehicleName": vehicle,
        "TotalCall": total_call,
        "InvestorCount": investor_count,
        "OutputFile": str(output_file),
    }])

    write_header = not log_file.exists()
    new_row.to_csv(log_file, mode="a", header=write_header, index=False)
    print(f"Run logged to: {log_file}")


# ================================
# PDF EXTRACTION FUNCTIONS
# ================================
def extract_text_from_pdf(pdf_path: Path) -> str:
    """Opens the PDF and pulls out all raw text, page by page."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_investment_name(text: str) -> str:
    """
    Pass 1: looks for legal entity suffixes (, LP / , LLC / , Ltd etc.)
    Pass 2: falls back to fund keyword scanning.
    """
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    skip_keywords = ["notice", "call", "date", "dear", "pursuant", "payment",
                     "wire", "bank", "suite", "street", "avenue", "floor", "email"]
    entity_pattern = re.compile(r",\s*(LP|LLC|Ltd\.?|L\.P\.|L\.L\.C\.|Inc\.?|Corp\.?)$", re.IGNORECASE)

    for line in lines[:25]:
        if any(skip in line.lower() for skip in skip_keywords):
            continue
        if entity_pattern.search(line):
            return line

    fund_keywords = ["fund", "capital", "partners", "ventures", "equity", "investments"]
    for line in lines[:25]:
        if any(skip in line.lower() for skip in skip_keywords):
            continue
        if any(kw in line.lower() for kw in fund_keywords):
            return line

    return "Unknown Investment"


def extract_due_date(text: str) -> str:
    """Extracts the due date, not the notice date."""
    match = re.search(r"Due\s+Date[:\s]+([A-Za-z]+\s+\d{1,2},?\s+\d{4})", text, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    match = re.search(r"Due\s+Date[:\s]+([\d]{1,2}[/\-][\d]{1,2}[/\-][\d]{2,4})", text, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    match = re.search(r"(?<!Due\s)Date\s*:\s*(.*)", text, re.IGNORECASE)
    return match.group(1).strip() if match else "Not Found"


def extract_total_capital_call(text: str) -> Optional[float]:
    """Finds the total capital call amount, anchored on $ to avoid matching percentages."""
    match = re.search(r"Total Capital Call[^$\n]*\$\s*([\d,]+\.\d{2})", text, re.IGNORECASE)
    if match:
        return float(match.group(1).replace(",", ""))
    match = re.search(r"Total Capital Call[^$\n]*\$\s*([\d,]+)(?!\.\d)", text, re.IGNORECASE)
    if match:
        return float(match.group(1).replace(",", ""))
    match = re.search(r"Total Capital Call[:\s]+([\d]{1,3}(?:,[\d]{3})+(?:\.\d{2})?)", text, re.IGNORECASE)
    if match:
        return float(match.group(1).replace(",", ""))
    return None


# ================================
# LOAD FROM SUPABASE
# ================================
def load_mapping(client) -> pd.DataFrame:
    """
    Queries the investment_mapping table in Supabase via the REST API.
    Returns a DataFrame with InvestmentName -> VehicleName mappings.
    """
    print("  Querying investment_mapping table...")
    rows = fetch_table(client, "investment_mapping")

    if not rows:
        raise ValueError(
            "investment_mapping table is empty or could not be reached.\n"
            "Check your Supabase table name and that data has been imported."
        )

    df = pd.DataFrame(rows)
    df["InvestmentName"] = df["InvestmentName"].astype(str).str.strip()
    df["VehicleName"] = df["VehicleName"].astype(str).str.strip()
    return df


def load_vehicle_investors(client) -> pd.DataFrame:
    """
    Queries the vehicle_investors table in Supabase via the REST API.
    Returns a DataFrame with the LP ownership splits per vehicle.
    """
    print("  Querying vehicle_investors table...")
    rows = fetch_table(client, "vehicle_investors")

    if not rows:
        raise ValueError(
            "vehicle_investors table is empty or could not be reached.\n"
            "Check your Supabase table name and that data has been imported."
        )

    df = pd.DataFrame(rows)

    # Keep only the three columns we need — ignore the id column Supabase adds
    df = df[["VehicleName", "Investor", "OwnershipPct"]].copy()
    df = df.dropna(subset=["VehicleName", "Investor", "OwnershipPct"])
    df["VehicleName"] = df["VehicleName"].astype(str).str.strip()
    df["Investor"] = df["Investor"].astype(str).str.strip()
    df["OwnershipPct"] = pd.to_numeric(df["OwnershipPct"], errors="coerce")

    if df["OwnershipPct"].isna().any():
        print("\nInvalid OwnershipPct rows found:")
        print(df[df["OwnershipPct"].isna()])
        raise ValueError("Fix the OwnershipPct values in Supabase, then re-run.")

    # If stored as whole numbers (e.g. 12.5 instead of 0.125), convert to decimal
    if df["OwnershipPct"].max() > 1:
        df["OwnershipPct"] = df["OwnershipPct"] / 100

    return df


# ================================
# BUSINESS LOGIC
# ================================
def find_vehicle_for_investment(investment_name: str, mapping_df: pd.DataFrame) -> str:
    """
    Matches the PDF fund name to a vehicle, checking both directions
    so partial matches work either way.
    """
    investment_clean = investment_name.lower().strip()

    for _, row in mapping_df.iterrows():
        db_name = row["InvestmentName"].lower().strip()
        if db_name in investment_clean or investment_clean in db_name:
            return row["VehicleName"]

    raise ValueError(
        f"No vehicle mapping found for: '{investment_name}'\n"
        f"Check the investment_mapping table in Supabase.\n"
        f"Available mappings:\n" +
        "\n".join(f"  - {r['InvestmentName']}" for _, r in mapping_df.iterrows())
    )


def get_investors_for_vehicle(vehicle_name: str, investors_df: pd.DataFrame) -> pd.DataFrame:
    """
    Filters to just the LPs for a specific vehicle.
    Normalizes ownership percentages and warns if they don't sum to 100%.
    """
    df = investors_df.loc[
        investors_df["VehicleName"].str.lower().str.strip() == vehicle_name.lower().strip()
    ].copy()

    if df.empty:
        raise ValueError(
            f"No investors found for vehicle: '{vehicle_name}'\n"
            f"Check the vehicle_investors table in Supabase.\n"
            f"Available vehicles:\n" +
            "\n".join(f"  - {v}" for v in investors_df["VehicleName"].unique())
        )

    total_pct = df["OwnershipPct"].sum()
    if not (0.999 < total_pct < 1.001):
        print(f"\nWARNING: Ownership % for '{vehicle_name}' sums to {total_pct:.4%}, not 100%.")
        print(f"         Normalizing automatically, but please fix the vehicle_investors table.\n")

    df["OwnershipPct"] = df["OwnershipPct"] / total_pct
    return df


# ================================
# ALLOCATION LOGIC
# ================================
def allocate_capital_call(total_call: float, investor_df: pd.DataFrame) -> pd.DataFrame:
    """
    Multiplies each LP's ownership % by the total call amount.
    Rounds to 2 decimal places and applies any rounding difference
    to the largest LP — standard fund accounting practice.
    """
    df = investor_df.copy()
    df["Allocation"] = (df["OwnershipPct"] * total_call).round(2)

    diff = round(total_call - df["Allocation"].sum(), 2)
    if diff != 0:
        idx = df["Allocation"].idxmax()
        df.loc[idx, "Allocation"] += diff
        print(f"Rounding adjustment of ${diff:.2f} applied to largest LP.")

    return df


# ================================
# OUTPUT FORMATTING (EXCEL)
# ================================
def format_excel(file_path: Path, investment_name: str, vehicle_name: str,
                 due_date: str, total_call: float):
    """
    Applies professional formatting to the output Excel file:
    title block, colored headers, number formats, auto-widths, totals row.
    """
    wb = load_workbook(file_path)
    ws = wb["Allocations"]

    ws.insert_rows(1, amount=5)

    ws["A1"] = "Capital Call Allocation Workpaper"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Investment:"
    ws["B2"] = investment_name
    ws["A2"].font = Font(bold=True)
    ws["A3"] = "Vehicle:"
    ws["B3"] = vehicle_name
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Due Date:"
    ws["B4"] = due_date
    ws["A4"].font = Font(bold=True)
    ws["A5"] = "Total Capital Call:"
    ws["B5"] = total_call
    ws["B5"].number_format = '"$"#,##0.00'
    ws["A5"].font = Font(bold=True)

    header_fill = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    last_data_row = ws.max_row
    header_row = {cell.value: cell.column for cell in ws[6] if cell.value}
    pct_col = header_row.get("OwnershipPct")
    alloc_col = header_row.get("Allocation")

    for row in ws.iter_rows(min_row=7, max_row=last_data_row):
        if pct_col:
            row[pct_col - 1].number_format = "0.00%"
        if alloc_col:
            row[alloc_col - 1].number_format = '"$"#,##0.00'

    totals_row = last_data_row + 1
    ws.cell(row=totals_row, column=1).value = "TOTAL"
    ws.cell(row=totals_row, column=1).font = Font(bold=True)

    if alloc_col:
        alloc_letter = get_column_letter(alloc_col)
        ws.cell(row=totals_row, column=alloc_col).value = f"=SUM({alloc_letter}7:{alloc_letter}{last_data_row})"
        ws.cell(row=totals_row, column=alloc_col).number_format = '"$"#,##0.00'
        ws.cell(row=totals_row, column=alloc_col).font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    wb.save(file_path)


# ================================
# MAIN EXECUTION
# ================================
def main():
    os.system("cls" if os.name == "nt" else "clear")

    print("=" * 50)
    print("  Capital Call Allocator")
    print("=" * 50)
    print()

    pdf_path = DATA_DIR / "sample.pdf"
    output_file = OUTPUT_DIR / "allocations.xlsx"

    print(f"Looking for PDF in: {DATA_DIR}")
    print()

    # Check PDF — no longer checking xlsx files, data comes from Supabase now
    print("Checking input files...")
    if not pdf_path.exists():
        print(f"  [MISSING] {pdf_path.name}")
        print("\nStopping — add the PDF then re-run.")
        return
    print(f"  [OK] {pdf_path.name}")
    print()

    # Connect to Supabase — stops here with a clear message if .env is missing
    print("Connecting to Supabase...")
    supabase = get_supabase_client()
    print("  Connected.")
    print()

    print("Step 1: Extracting data from PDF...")
    text = extract_text_from_pdf(pdf_path)
    investment_name = extract_investment_name(text)
    due_date = extract_due_date(text)
    total_call = extract_total_capital_call(text)

    print(f"  Investment : {investment_name}")
    print(f"  Due date   : {due_date}")
    print(f"  Total call : {total_call}")
    print()

    if total_call is None:
        print("Could not find 'Total Capital Call' amount in the PDF.")
        print("Check that the PDF contains a line like:")
        print("  Total Capital Call   $2,750,000.00")
        return

    print("Step 2: Loading data from Supabase...")
    mapping_df = load_mapping(supabase)
    investors_df = load_vehicle_investors(supabase)
    print(f"  Mapping rows  : {len(mapping_df)}")
    print(f"  Investor rows : {len(investors_df)}")
    print()

    print("Step 3: Mapping investment to vehicle...")
    vehicle_name = find_vehicle_for_investment(investment_name, mapping_df)
    print(f"  Mapped to: {vehicle_name}")
    print()

    print("Step 4: Loading investors for vehicle...")
    vehicle_investors = get_investors_for_vehicle(vehicle_name, investors_df)
    print(f"  Investors found: {len(vehicle_investors)}")
    print()

    print("Step 5: Calculating allocations...")
    allocation_df = allocate_capital_call(total_call, vehicle_investors)
    print()

    print("Step 6: Writing Excel output...")
    OUTPUT_DIR.mkdir(exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        allocation_df.to_excel(writer, sheet_name="Allocations", index=False)

    format_excel(output_file, investment_name, vehicle_name, due_date, total_call)
    print(f"  Saved to: {output_file}")
    print()

    print("Step 7: Logging run...")
    log_run(investment_name, vehicle_name, total_call, len(allocation_df), output_file)
    print()

    print("=" * 50)
    print("  SUCCESS")
    print("=" * 50)
    print(f"  Investment  : {investment_name}")
    print(f"  Vehicle     : {vehicle_name}")
    print(f"  Total call  : ${total_call:,.2f}")
    print(f"  Investors   : {len(allocation_df)}")
    print(f"  Output      : {output_file}")
    print()


# ================================
# RUN SCRIPT
# ================================
if __name__ == "__main__":
    main()