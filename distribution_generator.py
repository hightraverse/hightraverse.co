# ================================
# IMPORTS
# ================================
# os: clears terminal on startup
# re: regex for extracting fields from the distribution PDF
# Optional: type hint for values that might be None
# Path: cleaner file path handling than plain strings
# pd: pandas — reads Excel, handles LP data tables
# pdfplumber: opens PDF files and extracts raw text
# requests: hits the Supabase REST API
# load_dotenv: reads SUPABASE_URL and SUPABASE_KEY from .env
# load_workbook / Font etc.: opens and formats the Excel summary output
# get_column_letter: converts column numbers to letters dynamically
# reportlab: generates the branded distribution notice PDFs
import os
import re
from typing import Optional
from pathlib import Path

import pandas as pd
import pdfplumber
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
)


# ================================
# PATH SETUP
# ================================
BASE_DIR   = Path(__file__).resolve().parent
DATA_DIR   = BASE_DIR / "capital_calls" / "data"
OUTPUT_DIR = BASE_DIR / "capital_calls" / "output"
LOG_DIR    = BASE_DIR / "capital_calls" / "logs"
NOTICES_DIR = OUTPUT_DIR / "distribution_notices"
PDF_PATH   = DATA_DIR / "distribution_sample.pdf"


# ================================
# FIRM CONSTANTS
# ================================
FIRM_NAME    = "HighTraverse Capital Operations"
FIRM_EMAIL   = "info@hightraverse.co"
FIRM_WEBSITE = "hightraverse.co"


# ================================
# SUPABASE CONNECTION
# ================================
def get_supabase_client():
    """
    Loads credentials from .env and returns a (base_url, headers) tuple
    used for all Supabase REST API calls.
    """
    load_dotenv(BASE_DIR / ".env")
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        raise ValueError("SUPABASE_URL or SUPABASE_KEY not found in .env")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
    }
    return url.rstrip("/"), headers


def fetch_table(client, table_name: str) -> list:
    """
    Fetches every row from a Supabase table via REST API.
    Returns a list of dicts — one dict per row.
    """
    base_url, headers = client
    response = requests.get(
        f"{base_url}/rest/v1/{table_name}",
        headers=headers,
        params={"select": "*"},
        timeout=15,
    )
    response.raise_for_status()
    return response.json()


# ================================
# PDF EXTRACTION
# ================================
def extract_text_from_pdf(pdf_path: Path) -> str:
    """Opens the PDF and concatenates raw text from every page."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_investment_name(text: str) -> str:
    """
    Same two-pass logic as allocator.py:
    Pass 1 — look for legal entity suffixes (, LP / , LLC etc.)
    Pass 2 — fall back to fund keyword scanning.
    """
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    skip_keywords = ["notice", "distribution", "date", "dear", "pursuant",
                     "payment", "wire", "bank", "suite", "street", "avenue",
                     "floor", "email", "pleased"]
    entity_pattern = re.compile(
        r",\s*(LP|LLC|Ltd\.?|L\.P\.|L\.L\.C\.|Inc\.?|Corp\.?)$", re.IGNORECASE
    )

    for line in lines[:25]:
        if any(skip in line.lower() for skip in skip_keywords):
            continue
        if entity_pattern.search(line):
            return line

    fund_keywords = ["fund", "capital", "partners", "ventures", "equity", "investments"]
    for line in lines[:25]:
        if any(skip in line.lower() for skip in skip_keywords):
            continue
        if any(kw in line.lower() for kw in fund_keywords):
            return line

    return "Unknown Investment"


def extract_payment_date(text: str) -> str:
    """
    Looks for a Payment Date line first (the date proceeds actually go out),
    then falls back to a general Date line.
    """
    match = re.search(
        r"Payment\s+Date[:\s]+([A-Za-z]+\s+\d{1,2},?\s+\d{4})", text, re.IGNORECASE
    )
    if match:
        return match.group(1).strip()
    match = re.search(
        r"Payment\s+Date[:\s]+([\d]{1,2}[/\-][\d]{1,2}[/\-][\d]{2,4})",
        text, re.IGNORECASE
    )
    if match:
        return match.group(1).strip()
    match = re.search(r"Date\s*:\s*(.*)", text, re.IGNORECASE)
    return match.group(1).strip() if match else "Not Found"


def extract_total_distribution(text: str) -> Optional[float]:
    """
    Finds the total distribution amount. Anchored on 'Total Distribution'
    so we don't accidentally pick up other dollar figures on the page.
    """
    match = re.search(
        r"Total Distribution[^$\n]*\$\s*([\d,]+\.\d{2})", text, re.IGNORECASE
    )
    if match:
        return float(match.group(1).replace(",", ""))
    match = re.search(
        r"Total Distribution[^$\n]*\$\s*([\d,]+)(?!\.\d)", text, re.IGNORECASE
    )
    if match:
        return float(match.group(1).replace(",", ""))
    match = re.search(
        r"Total Distribution[:\s]+([\d]{1,3}(?:,[\d]{3})+(?:\.\d{2})?)",
        text, re.IGNORECASE
    )
    if match:
        return float(match.group(1).replace(",", ""))
    return None


# ================================
# LOAD FROM SUPABASE
# ================================
def load_mapping(client) -> pd.DataFrame:
    """Fetches investment_mapping — InvestmentName → VehicleName."""
    print("  Querying investment_mapping table...")
    rows = fetch_table(client, "investment_mapping")
    if not rows:
        raise ValueError("investment_mapping table is empty or unreachable.")
    df = pd.DataFrame(rows)
    df["InvestmentName"] = df["InvestmentName"].astype(str).str.strip()
    df["VehicleName"]    = df["VehicleName"].astype(str).str.strip()
    return df


def load_vehicle_investors(client) -> pd.DataFrame:
    """Fetches vehicle_investors — ownership splits per vehicle."""
    print("  Querying vehicle_investors table...")
    rows = fetch_table(client, "vehicle_investors")
    if not rows:
        raise ValueError("vehicle_investors table is empty or unreachable.")
    df = pd.DataFrame(rows)
    df = df[["VehicleName", "Investor", "OwnershipPct"]].copy()
    df = df.dropna(subset=["VehicleName", "Investor", "OwnershipPct"])
    df["VehicleName"]  = df["VehicleName"].astype(str).str.strip()
    df["Investor"]     = df["Investor"].astype(str).str.strip()
    df["OwnershipPct"] = pd.to_numeric(df["OwnershipPct"], errors="coerce")
    if df["OwnershipPct"].max() > 1:
        df["OwnershipPct"] = df["OwnershipPct"] / 100
    return df


# ================================
# BUSINESS LOGIC
# ================================
def find_vehicle_for_investment(investment_name: str, mapping_df: pd.DataFrame) -> str:
    """Maps the PDF fund name to the correct fund-of-funds vehicle."""
    investment_clean = investment_name.lower().strip()
    for _, row in mapping_df.iterrows():
        db_name = row["InvestmentName"].lower().strip()
        if db_name in investment_clean or investment_clean in db_name:
            return row["VehicleName"]
    raise ValueError(
        f"No vehicle mapping found for: '{investment_name}'\n"
        "Check investment_mapping in Supabase.\nAvailable:\n" +
        "\n".join(f"  - {r['InvestmentName']}" for _, r in mapping_df.iterrows())
    )


def get_investors_for_vehicle(vehicle_name: str, investors_df: pd.DataFrame) -> pd.DataFrame:
    """Filters to just the LPs for a specific vehicle and normalizes ownership %."""
    df = investors_df.loc[
        investors_df["VehicleName"].str.lower().str.strip() == vehicle_name.lower().strip()
    ].copy()
    if df.empty:
        raise ValueError(f"No investors found for vehicle: '{vehicle_name}'")
    total_pct = df["OwnershipPct"].sum()
    if not (0.999 < total_pct < 1.001):
        print(f"\nWARNING: Ownership % for '{vehicle_name}' sums to {total_pct:.4%}, not 100%.")
        print("         Normalizing automatically.\n")
    df["OwnershipPct"] = df["OwnershipPct"] / total_pct
    return df


def allocate_distribution(total_distribution: float, investor_df: pd.DataFrame) -> pd.DataFrame:
    """
    Multiplies each LP's ownership % by the total distribution amount.
    Applies any rounding difference to the largest LP — same approach as
    capital call allocation, standard fund accounting practice.
    """
    df = investor_df.copy()
    df["Distribution"] = (df["OwnershipPct"] * total_distribution).round(2)
    diff = round(total_distribution - df["Distribution"].sum(), 2)
    if diff != 0:
        idx = df["Distribution"].idxmax()
        df.loc[idx, "Distribution"] += diff
        print(f"  Rounding adjustment of ${diff:.2f} applied to largest LP.")
    return df


# ================================
# EXCEL SUMMARY OUTPUT
# ================================
def write_excel_summary(df: pd.DataFrame, investment_name: str, vehicle_name: str,
                         payment_date: str, total_distribution: float,
                         output_path: Path):
    """
    Writes a formatted Excel summary of the distribution allocations —
    same header block + formatting approach as allocator.py's output.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Distributions", index=False)

    wb = load_workbook(output_path)
    ws = wb["Distributions"]
    ws.insert_rows(1, amount=5)

    ws["A1"] = "Distribution Allocation Workpaper"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Investment:"
    ws["B2"] = investment_name
    ws["A2"].font = Font(bold=True)
    ws["A3"] = "Vehicle:"
    ws["B3"] = vehicle_name
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "Payment Date:"
    ws["B4"] = payment_date
    ws["A4"].font = Font(bold=True)
    ws["A5"] = "Total Distribution:"
    ws["B5"] = total_distribution
    ws["B5"].number_format = '"$"#,##0.00'
    ws["A5"].font = Font(bold=True)

    header_fill = PatternFill(start_color="D5F0E0", end_color="D5F0E0", fill_type="solid")
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    last_data_row = ws.max_row
    header_row = {cell.value: cell.column for cell in ws[6] if cell.value}
    pct_col   = header_row.get("OwnershipPct")
    dist_col  = header_row.get("Distribution")

    for row in ws.iter_rows(min_row=7, max_row=last_data_row):
        if pct_col:
            row[pct_col - 1].number_format = "0.00%"
        if dist_col:
            row[dist_col - 1].number_format = '"$"#,##0.00'

    totals_row = last_data_row + 1
    ws.cell(row=totals_row, column=1).value = "TOTAL"
    ws.cell(row=totals_row, column=1).font = Font(bold=True)
    if dist_col:
        dist_letter = get_column_letter(dist_col)
        ws.cell(row=totals_row, column=dist_col).value = \
            f"=SUM({dist_letter}7:{dist_letter}{last_data_row})"
        ws.cell(row=totals_row, column=dist_col).number_format = '"$"#,##0.00'
        ws.cell(row=totals_row, column=dist_col).font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    wb.save(output_path)


# ================================
# GENERATE ONE DISTRIBUTION NOTICE PDF
# ================================
def build_styles():
    """Returns named ParagraphStyles for the distribution notice PDFs."""
    base = getSampleStyleSheet()
    navy  = colors.HexColor("#1a2e4a")
    grey  = colors.HexColor("#555555")
    green = colors.HexColor("#1e6b3a")

    return {
        "firm_name": ParagraphStyle("firm_name", parent=base["Normal"],
            fontSize=16, fontName="Helvetica-Bold", textColor=navy, spaceAfter=2),
        "firm_contact": ParagraphStyle("firm_contact", parent=base["Normal"],
            fontSize=9, fontName="Helvetica", textColor=grey, spaceAfter=0),
        "doc_type": ParagraphStyle("doc_type", parent=base["Normal"],
            fontSize=10, fontName="Helvetica-Bold", textColor=green,
            spaceBefore=18, spaceAfter=4),
        "label": ParagraphStyle("label", parent=base["Normal"],
            fontSize=10, fontName="Helvetica-Bold", textColor=colors.HexColor("#333333"),
            spaceAfter=2),
        "body": ParagraphStyle("body", parent=base["Normal"],
            fontSize=10, fontName="Helvetica", textColor=colors.HexColor("#333333"),
            leading=15, spaceAfter=8),
        "amount_label": ParagraphStyle("amount_label", parent=base["Normal"],
            fontSize=10, fontName="Helvetica", textColor=grey, spaceAfter=2),
        "amount_value": ParagraphStyle("amount_value", parent=base["Normal"],
            fontSize=22, fontName="Helvetica-Bold", textColor=green, spaceAfter=6),
        "footer": ParagraphStyle("footer", parent=base["Normal"],
            fontSize=8, fontName="Helvetica", textColor=grey),
    }


def generate_notice(investor_name: str, ownership_pct: float, distribution: float,
                    investment_name: str, vehicle_name: str, payment_date: str,
                    total_distribution: float, styles: dict, output_path: Path):
    """
    Generates one distribution notice PDF for a single LP.

    Layout mirrors the capital call memo but with distribution language:
    - Green accent color instead of navy (signals money coming IN to the LP)
    - "You will receive" framing rather than "you owe"
    - No wire instructions section — the vehicle is sending to the LP,
      so the LP doesn't need to do anything except await the wire
    """
    from datetime import date as dt
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        leftMargin=0.9 * inch, rightMargin=0.9 * inch,
        topMargin=0.8 * inch,  bottomMargin=0.8 * inch,
    )
    story = []

    # Letterhead
    story.append(Paragraph(FIRM_NAME, styles["firm_name"]))
    story.append(Paragraph(f"{FIRM_EMAIL}  ·  {FIRM_WEBSITE}", styles["firm_contact"]))
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=2,
        color=colors.HexColor("#1e6b3a"), spaceAfter=0))

    # Document type
    story.append(Paragraph("DISTRIBUTION NOTICE", styles["doc_type"]))

    # Memo header table
    today = dt.today().strftime("%B %d, %Y")
    header_data = [
        [Paragraph("<b>TO:</b>",   styles["body"]), Paragraph(investor_name, styles["body"])],
        [Paragraph("<b>FROM:</b>", styles["body"]), Paragraph(FIRM_NAME,     styles["body"])],
        [Paragraph("<b>DATE:</b>", styles["body"]), Paragraph(today,          styles["body"])],
        [Paragraph("<b>RE:</b>",   styles["body"]), Paragraph(
            f"Distribution — {investment_name} / {vehicle_name}", styles["body"])],
    ]
    header_table = Table(header_data, colWidths=[0.8 * inch, 5.8 * inch])
    header_table.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(header_table)
    story.append(HRFlowable(width="100%", thickness=0.5,
        color=colors.HexColor("#cccccc"), spaceBefore=4, spaceAfter=14))

    # Body paragraph
    body_text = (
        f"We are pleased to advise that {investment_name} (the \"Fund\") has declared a "
        f"distribution to {vehicle_name} (the \"Vehicle\"). Pursuant to your limited "
        f"partnership interest in the Vehicle, your pro-rata distribution is set forth below."
        f"<br/><br/>"
        f"Distribution proceeds are expected to be wired to the bank account on file no later "
        f"than <b>{payment_date}</b>. Please contact {FIRM_EMAIL} if you have not received "
        f"payment by that date or if your wire instructions have changed."
    )
    story.append(Paragraph(body_text, styles["body"]))

    # Distribution amount — green to signal money coming in
    story.append(Spacer(1, 6))
    story.append(Paragraph("YOUR DISTRIBUTION AMOUNT", styles["amount_label"]))
    story.append(Paragraph(f"${distribution:,.2f}", styles["amount_value"]))
    story.append(HRFlowable(width="100%", thickness=0.5,
        color=colors.HexColor("#cccccc"), spaceBefore=2, spaceAfter=14))

    # Details table
    details_data = [
        ["Fund / Investment",    investment_name],
        ["Vehicle",              vehicle_name],
        ["Total Distribution",   f"${total_distribution:,.2f}"],
        ["Your Ownership %",     f"{ownership_pct:.4%}"],
        ["Your Distribution",    f"${distribution:,.2f}"],
        ["Expected Payment Date", str(payment_date)],
    ]
    details_table = Table(details_data, colWidths=[2.2 * inch, 4.4 * inch])
    details_table.setStyle(TableStyle([
        ("FONTNAME",      (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("TEXTCOLOR",     (0, 0), (0, -1), colors.HexColor("#333333")),
        ("TEXTCOLOR",     (1, 0), (1, -1), colors.HexColor("#555555")),
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#f0f8f0")),
        ("BACKGROUND",    (0, 2), (-1, 2), colors.HexColor("#f0f8f0")),
        ("BACKGROUND",    (0, 4), (-1, 4), colors.HexColor("#d5f0e0")),
        ("FONTNAME",      (0, 4), (-1, 4), "Helvetica-Bold"),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
    ]))
    story.append(details_table)

    # Footer
    story.append(Spacer(1, 30))
    story.append(HRFlowable(width="100%", thickness=0.5,
        color=colors.HexColor("#cccccc"), spaceAfter=6))
    story.append(Paragraph(
        f"This notice has been prepared by {FIRM_NAME} on behalf of {vehicle_name}. "
        f"Confidential — intended solely for the named recipient. "
        f"Questions: {FIRM_EMAIL}",
        styles["footer"],
    ))

    doc.build(story)


# ================================
# LOGGING
# ================================
def log_run(investment: str, vehicle: str, total_dist: float,
            investor_count: int, output_file: Path):
    """Appends one row to logs/run_log.csv for audit trail."""
    LOG_DIR.mkdir(exist_ok=True)
    log_file = LOG_DIR / "run_log.csv"
    new_row = pd.DataFrame([{
        "Timestamp":     pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Type":          "Distribution",
        "InvestmentName": investment,
        "VehicleName":   vehicle,
        "TotalAmount":   total_dist,
        "InvestorCount": investor_count,
        "OutputFile":    str(output_file),
    }])
    write_header = not log_file.exists()
    new_row.to_csv(log_file, mode="a", header=write_header, index=False)
    print(f"  Run logged to: {log_file}")


# ================================
# MAIN EXECUTION
# ================================
def main():
    os.system("cls" if os.name == "nt" else "clear")

    print("=" * 50)
    print("  Distribution Notice Generator")
    print("=" * 50)
    print()

    # Check PDF exists
    print("Checking input files...")
    if not PDF_PATH.exists():
        print(f"  [MISSING] {PDF_PATH.name}")
        print("\nStopping — add distribution_sample.pdf to capital_calls/data/ and re-run.")
        return
    print(f"  [OK] {PDF_PATH.name}")
    print()

    # Connect to Supabase
    print("Connecting to Supabase...")
    client = get_supabase_client()
    print("  Connected.")
    print()

    # Step 1: Extract from PDF
    print("Step 1: Extracting data from PDF...")
    text = extract_text_from_pdf(PDF_PATH)
    investment_name    = extract_investment_name(text)
    payment_date       = extract_payment_date(text)
    total_distribution = extract_total_distribution(text)

    print(f"  Investment      : {investment_name}")
    print(f"  Payment date    : {payment_date}")
    print(f"  Total distribution : {total_distribution}")
    print()

    if total_distribution is None:
        print("Could not find 'Total Distribution' amount in the PDF.")
        print("Check the PDF contains a line like:")
        print("  Total Distribution   $1,250,000.00")
        return

    # Step 2: Load Supabase data
    print("Step 2: Loading data from Supabase...")
    mapping_df   = load_mapping(client)
    investors_df = load_vehicle_investors(client)
    print(f"  Mapping rows  : {len(mapping_df)}")
    print(f"  Investor rows : {len(investors_df)}")
    print()

    # Step 3: Map investment → vehicle
    print("Step 3: Mapping investment to vehicle...")
    vehicle_name = find_vehicle_for_investment(investment_name, mapping_df)
    print(f"  Mapped to: {vehicle_name}")
    print()

    # Step 4: Get investors for vehicle
    print("Step 4: Loading investors for vehicle...")
    vehicle_investors = get_investors_for_vehicle(vehicle_name, investors_df)
    print(f"  Investors found: {len(vehicle_investors)}")
    print()

    # Step 5: Calculate distributions
    print("Step 5: Calculating distributions...")
    distribution_df = allocate_distribution(total_distribution, vehicle_investors)
    print()

    # Step 6: Write Excel summary
    print("Step 6: Writing Excel summary...")
    OUTPUT_DIR.mkdir(exist_ok=True)
    excel_path = OUTPUT_DIR / "distributions.xlsx"
    write_excel_summary(
        distribution_df, investment_name, vehicle_name,
        payment_date, total_distribution, excel_path
    )
    print(f"  Saved to: {excel_path}")
    print()

    # Step 7: Generate one PDF notice per LP
    print("Step 7: Generating distribution notices...")
    NOTICES_DIR.mkdir(parents=True, exist_ok=True)
    styles = build_styles()
    generated = []
    errors    = []

    for _, row in distribution_df.iterrows():
        investor_name = row["Investor"]
        ownership_pct = row["OwnershipPct"]
        distribution  = row["Distribution"]

        safe_name = "".join(
            c if c.isalnum() or c in (" ", "-") else "_"
            for c in investor_name
        ).strip()
        output_path = NOTICES_DIR / f"{safe_name} - Distribution Notice.pdf"

        try:
            generate_notice(
                investor_name, ownership_pct, distribution,
                investment_name, vehicle_name, payment_date,
                total_distribution, styles, output_path
            )
            print(f"  [OK] {output_path.name}")
            generated.append(output_path)
        except Exception as e:
            print(f"  [ERROR] {investor_name}: {e}")
            errors.append((investor_name, str(e)))

    print()

    # Step 8: Log run
    print("Step 8: Logging run...")
    log_run(investment_name, vehicle_name, total_distribution,
            len(distribution_df), excel_path)
    print()

    # Summary
    print("=" * 50)
    print("  SUCCESS")
    print("=" * 50)
    print(f"  Investment    : {investment_name}")
    print(f"  Vehicle       : {vehicle_name}")
    print(f"  Total dist.   : ${total_distribution:,.2f}")
    print(f"  Investors     : {len(distribution_df)}")
    print(f"  Notices saved : {len(generated)}")
    if errors:
        print(f"  Errors        : {len(errors)}")
    print(f"  Excel summary : {excel_path}")
    print(f"  Notices folder: {NOTICES_DIR}")
    print()


# ================================
# RUN SCRIPT
# ================================
if __name__ == "__main__":
    main()
